from flask import Flask, render_template, request, redirect, url_for, session
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

app = Flask(__name__)
app.secret_key = "my_secret_key_for_learning"

EXCEL_FILE = "users.xlsx"


def create_excel_file():
    if not os.path.exists(EXCEL_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Users"

        sheet.append([
            "Email",
            "Password",
            "Daily Note",
            "Submitted At"
        ])

        workbook.save(EXCEL_FILE)


def user_exists(email):
    create_excel_file()

    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        saved_email = row[0]

        if saved_email == email:
            return True

    return False


def check_login(email, password):
    create_excel_file()

    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        saved_email = row[0]
        saved_password = row[1]

        if saved_email == email and saved_password == password:
            return True

    return False


def get_user_notes():
    create_excel_file()

    email = session.get("email")

    if not email:
        return []

    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active

    notes = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        saved_email = row[0]
        daily_note = row[2]
        submitted_at = row[3]

        if saved_email == email and daily_note != "Account created":
            notes.append({
                "email": saved_email,
                "note": daily_note,
                "submitted_at": submitted_at
            })

    return notes


@app.route("/")
def login():
    return render_template("login.html")


@app.route("/home")
def home():
    if "email" not in session:
        return redirect(url_for("login"))

    return render_template("home.html")


@app.route("/dashboard")
def dashboard():
    if "email" not in session:
        return redirect(url_for("login"))

    return render_template("dashboard.html")


@app.route("/profile")
def profile():
    if "email" not in session:
        return redirect(url_for("login"))

    notes = get_user_notes()
    return render_template("profile.html", notes=notes)


@app.route("/login", methods=["POST"])
def login_user():
    email = request.form.get("email")
    password = request.form.get("password")

    if check_login(email, password):
        session["email"] = email
        session["password"] = password
        return redirect(url_for("home"))
    else:
        return render_template(
            "login.html",
            error="Incorrect email or password. Please try again."
        )


@app.route("/signup", methods=["POST"])
def signup_user():
    email = request.form.get("email")
    password = request.form.get("password")

    create_excel_file()

    if user_exists(email):
        return render_template(
            "login.html",
            error="This email already exists. Please login instead."
        )

    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active

    sheet.append([
        email,
        password,
        "Account created",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ])

    workbook.save(EXCEL_FILE)

    session["email"] = email
    session["password"] = password

    return redirect(url_for("home"))


@app.route("/submit_note", methods=["POST"])
def submit_note():
    if "email" not in session:
        return redirect(url_for("login"))

    daily_note = request.form.get("daily_note")

    email = session.get("email")
    password = session.get("password")

    create_excel_file()

    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active

    sheet.append([
        email,
        password,
        daily_note,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ])

    workbook.save(EXCEL_FILE)

    return redirect(url_for("profile"))


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


if __name__ == "__main__":
    app.run(debug=True)